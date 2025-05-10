import os
import zipfile
import base64
import glob
import json
import tempfile
import shutil
import pandas as pd 
import logging
from io import BytesIO
from dotenv import load_dotenv
from PIL import Image    
from pdf2image import convert_from_path   
from openai import OpenAI  
from openpyxl.styles import Font 
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import streamlit as st      
from google import genai
from google.genai.types import Part             
import ast
import pathlib

def check_auth():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    if not st.session_state.logged_in:
        st.subheader("🔒 Login to Access App")
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")

        valid_users = ast.literal_eval(os.getenv("APP_USERS", "{}"))

        if st.button("Login"):
            if username in valid_users and password == valid_users[username]:
                st.session_state.logged_in = True
                st.success("✅ Logged in successfully")
                st.rerun()
            else:
                st.error("❌ Invalid username or password")
        return False
    return True

# ---------- Load Keys ----------
load_dotenv()
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
genai_client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

vendor_categories = {
    "SA Imports": ["Rasheeda Industries", "L&K Poly"],
    "Clearing, FF, Duties, W/housing": ["Shuttle Freight"],
    "Meat": ["St Helens", "Sargent Farms", "Toronto Halal", "Sysco", "Solmaz", "Bacha Casings", "J&FC Seafood"],
    "Local Purchases": ["Sysco", "A1", "Bun Man", "Sahel Khan", "Starsky", "Mr Produce", "Walmart", "PIX Graphics", "Kim Eco Pak"],
    "Utilities": ["Alectra", "Enbridge", "Rogers", "Telus"],
    "Rent": ["SDEB"],
    "Repairs & Maintenance": ["MechArm", "IB Technical", "Just instruments", "Willy Oosthuizen", "Skypole"],
    "Office Expenses": ["Amazon"],
    "Health & Safety": ["Green Planet", "Abell Pest Control", "Cintas", "CFIA", "HMA", "Waste Connection of Canada", "Aqua Team Power Clean"],
    "Insurance, Legal, Accounting": ["Des Jardin", "HHAcc Services"],
    "Bank charges": ["RBC", "Moneris"],
    "Delivery service": ["Uber Eats", "Door Dash"],
    "Freight on line shopping": ["Click ship (Freight.com)"],
    "Vehicle & Fuel": ["Nissan", "Stinton", "Shell"],
    "Consulting": ["Food safety first"]
}

receipt_schema = {
    "type": "object",
    "properties": {
        "vendor_name": {"type": "string"},
        "total_amount": {"type": "string"},
        "tax_total": {"type": "string"},
        "date": {"type": "string"},
        "category": {"type": "string"}
    },
    "required": ["vendor_name", "total_amount", "tax_total", "date", "category"],
    "additionalProperties": False
}


# ---------- Statement Parsing ----------
def extract_statement(pdf_path):
    filepath = pathlib.Path(pdf_path)
    prompt =  """Analyze the text in the provided document. Extract all readable content
            and present it in a structured Markdown format that is clear, concise, 
            and well-organized. Use headings, lists, or tables where appropriate."""
    response = genai_client.models.generate_content(
    model="gemini-2.0-flash",
    contents=[
        Part.from_bytes(
            data=filepath.read_bytes(),
            mime_type='application/pdf',
        ),
        prompt])
    print(response.text)
    return response.text

# ---------- OCR & Invoice Extraction ----------
def pdf_to_base64_images(pdf_path):
    images = convert_from_path(pdf_path, dpi=200)
    base64_images = []
    for i, image in enumerate(images):
        temp_path = f"temp_page_{i+1}.jpg"
        image.save(temp_path, "JPEG")
        with open(temp_path, "rb") as f:
            base64_images.append(base64.b64encode(f.read()).decode("utf-8"))
    return base64_images

def extract_invoice_info(pdf_path):
    base64_images = pdf_to_base64_images(pdf_path)
    message_content = [{
        "type": "text",
        "text": (
            f"""You're a receipt parser. The following images are pages from one receipt. Extract the vendor name, total amount, tax total/hst and date from the receipt image. 
            Return in JSON format.If Total amount is 0.00, extract subtotal and add on any tax and save it under Total. Always format date in MM/DD/YYYY format.
            Do not wrap output in ```json ```  
            
            Here are the vendor categories:
            {vendor_categories}
            
            Please assign the receipt to one of the above categories based on the vendor name and add this to the json return.
            """
        )
    }]
    for b64_img in base64_images:
        message_content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/jpeg;base64,{b64_img}"
            }
        })

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": message_content}],
        response_format={
            "type": "json_schema",
            "json_schema": {
                "name": "receipt_info",
                "schema": receipt_schema,
                "strict": True
            }
        },
        max_tokens=500
    )
    return response.choices[0].message.content

# ---------- Reconciliation ----------
def reconcile_with_statement(invoice_json, statement_md):
    prompt = f"""
- You are an invoice reconciler agent. You will be given a json of an invoice and your job is to reconcile the invoice with the credit card or bank statement.
- If the invoice is found in the statement, append to the input json with the following key value pair: "reconciled": true
- If the invoice is not found in the statement, append to the input json with the following key value pair: "reconciled": false
- Do not wrap output in ```json ```

Here is the invoice:
---
{invoice_json}
---
And here is the credit card or bank statement markdown:
---
{statement_md}
---
"""
    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=500
    )
    return response.choices[0].message.content

def uploaded_pdf_to_tempfile(uploaded_file):
    suffix = ".pdf"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(uploaded_file.read())
        return temp_file.name
    
def uploaded_zip_to_tempfile(uploaded_file):
    suffix = ".zip"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(uploaded_file.read())
        return temp_file.name


def unzip_and_process(zip_file):
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(zip_file, "r") as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def split_and_export(results, label):
    matched = [r for r in results if r.get("reconciled") is True]
    if not matched:
        return None
    df = pd.DataFrame(matched)
    output_path = f"reconciled_{label}.xlsx"
    df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    bold_font = Font(bold=True)
    for col_num in range(1, len(df.columns) + 1):
        ws[f"{get_column_letter(col_num)}1"].font = bold_font
    wb.save(output_path)
    return output_path


# ---------- Streamlit App ----------
if check_auth():
    st.set_page_config(page_title="Invoice Reconciler", layout="centered")
    st.title("📑 Invoice Reconciliation App")

    if st.button("Logout"):
        st.session_state.logged_in = False
        st.rerun()

    cc_pdf = st.file_uploader("Upload Credit Card Statement (PDF)", type="pdf")
    bank_pdf = st.file_uploader("Upload Bank Statement (PDF)", type="pdf")
    cc_zip = st.file_uploader("Upload Invoices for Credit Card (ZIP)", type="zip")
    bank_zip = st.file_uploader("Upload Invoices for Bank (ZIP)", type="zip")

    if st.button("🔄 Run Reconciliation") and cc_pdf and cc_zip and bank_zip and bank_pdf:
        print("starting")
        with st.spinner("🔍 Extracting statements..."):
            try:
                credit_path = uploaded_pdf_to_tempfile(cc_pdf)
                cc_md = extract_statement(credit_path)
                print("CC statement extracted")
            except Exception as e:
                st.error(f"Error extracting credit card statement: {e}")
            
            try:
                bank_md_path = uploaded_pdf_to_tempfile(bank_pdf)
                bank_md = extract_statement(bank_md_path)
                print("Bank statement extracted")
            except Exception as e:
                st.error(f"Error extracting bank statement: {e}")

            
            
        cc_folder = unzip_and_process(uploaded_zip_to_tempfile(cc_zip))
        bank_folder = unzip_and_process(uploaded_zip_to_tempfile(bank_zip))

        print("checkpoint")

        print(cc_folder)
        print(bank_folder)
        results_cc, results_bank = [], []

        if cc_md:
            with st.spinner("🤖 Processing Credit Card invoices..."):
                for file in glob.glob(cc_folder + '/*'):
                    for f in glob.glob(file + '/*'):
                        if f.endswith(".pdf"):
                            info = extract_invoice_info(f)
                            reconciled = reconcile_with_statement(info, cc_md)
                            results_cc.append(json.loads(reconciled))

        print(results_cc)

        if bank_md:
            with st.spinner("🤖 Processing Bank invoices..."):
                for file in glob.glob(bank_folder + '/*'):
                    for f in glob.glob(file + '/*'):
                        if f.endswith(".pdf"):
                            info = extract_invoice_info(f)
                            reconciled = reconcile_with_statement(info, bank_md)
                            results_bank.append(json.loads(reconciled))

        print(results_bank)

        cc_path = split_and_export(results_cc, "cc")
        bank_path = split_and_export(results_bank, "bank")

        if cc_path:
            with open(cc_path, "rb") as f:
                cc_bytes = f.read()
                st.session_state["cc_bytes"] = cc_bytes

        if bank_path:
            with open(bank_path, "rb") as f:
                bank_bytes = f.read()
                st.session_state["bank_bytes"] = bank_bytes
        st.success("✅ Reconciliation complete!")

    if "cc_bytes" in st.session_state:
        st.download_button("📥 Download Credit Card Reconciled", st.session_state["cc_bytes"], file_name="cc_reconciled.xlsx")

    if "bank_bytes" in st.session_state:
        st.download_button("📥 Download Bank Reconciled", st.session_state["bank_bytes"], file_name="bank_reconciled.xlsx")

