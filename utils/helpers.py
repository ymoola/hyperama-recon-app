import tempfile
import zipfile
import pandas as pd
from config.settings import openai_client
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from utils.prompts import reconciliation_prompt


def uploaded_pdf_to_tempfile(uploaded_file):
    suffix = ".pdf"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(uploaded_file.read())
        return temp_file.name

def uploaded_zip_to_tempfile(uploaded_file):
    suffix = ".zip"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(uploaded_file.read())
        return temp_file.name

def unzip_and_process(zip_file):
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(zip_file, "r") as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def reconcile_with_statement(invoice_json, statement_md):
    prompt = reconciliation_prompt.format(invoice_json, statement_md)
    response = openai_client.chat.completions.create(
        model= "gpt-4.1-mini",
        messages=[{"role": "user", "content": prompt}]
    )
    return response.choices[0].message.content


def export_combined_results(results_cc, results_bank, output_path="reconciliation_results.xlsx"):
    matched, unmatched = [], []

    for entry in results_cc:
        entry["source"] = "credit card"
        if entry.get("reconciled") is True:
            matched.append(entry)
        else:
            unmatched.append(entry)

    for entry in results_bank:
        entry["source"] = "bank"
        if entry.get("reconciled") is True:
            matched.append(entry)
        else:
            unmatched.append(entry)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if matched:
            matched_df = pd.DataFrame(matched)
            matched_df.to_excel(writer, sheet_name="Matched", index=False)
            ws = writer.sheets["Matched"]
            for col_num, column_title in enumerate(matched_df.columns, 1):
                ws[f"{get_column_letter(col_num)}1"].font = Font(bold=True)

        if unmatched:
            unmatched_df = pd.DataFrame(unmatched)
            unmatched_df.to_excel(writer, sheet_name="Unmatched", index=False)
            ws = writer.sheets["Unmatched"]
            for col_num, column_title in enumerate(unmatched_df.columns, 1):
                ws[f"{get_column_letter(col_num)}1"].font = Font(bold=True)

    return output_path

