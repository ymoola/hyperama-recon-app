import pathlib
import glob
import pandas as pd
import time
from collections import deque
from config.settings import genai_client, openai_client, COLUMN_GROUPS, SalesReport
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from google.genai.types import Part
from utils.prompts import sales_extraction_prompt



# Tracks timestamps of API calls
request_timestamps = deque()

# Rate limit constants
MAX_REQUESTS_PER_MIN = 9
WINDOW_SECONDS = 60

def enforce_rate_limit():
    now = time.time()
    
    # Remove timestamps older than 60 seconds
    while request_timestamps and now - request_timestamps[0] > WINDOW_SECONDS:
        request_timestamps.popleft()
    
    if len(request_timestamps) >= MAX_REQUESTS_PER_MIN:
        sleep_time = WINDOW_SECONDS - (now - request_timestamps[0])
        print(f"⏳ Rate limit reached. Sleeping for {sleep_time:.1f} seconds...")
        time.sleep(sleep_time)
        enforce_rate_limit()  # recheck after sleep


def extract_sales_data(pdf_path: str) -> dict:
    enforce_rate_limit()
    
    response = genai_client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=[
            Part.from_bytes(data=pathlib.Path(pdf_path).read_bytes(), mime_type='application/pdf'),
            sales_extraction_prompt
        ],
        config={
            "response_mime_type": "application/json",
            "response_schema": SalesReport
        }
    )
    
    request_timestamps.append(time.time())
    return response.parsed.model_dump()

def write_to_excel_with_categories(df: pd.DataFrame, output_excel: str):
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y")
    df.sort_values("Date", inplace=True)
    df["Date"] = df["Date"].dt.strftime("%d-%m-%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    all_fields = ["Date"]
    for fields in COLUMN_GROUPS.values():
        all_fields.extend(fields)
    all_fields.append("Notes")

    first_row = [""]  # Date column top
    merge_ranges = []
    col = 2
    for category, fields in COLUMN_GROUPS.items():
        first_row.extend([category] + [""] * (len(fields) - 1))
        merge_ranges.append((col, col + len(fields) - 1, category))
        col += len(fields)
    first_row.append("")

    ws.append(first_row)
    ws.append(all_fields)

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for start_col, end_col, category in merge_ranges:
        ws.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1")
        cell = ws[f"{get_column_letter(start_col)}1"]
        cell.value = category
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        fill_color = {
            "RETAIL SUPERMARKET & BUTCHERY": "FFD700",
            "DINER": "ADD8E6",
            "SHOPIFY": "98FB98",
            "SALES SUMMARY": "F08080"
        }.get(category, "DDDDDD")

        for col in range(start_col, end_col + 1):
            ws[f"{get_column_letter(col)}1"].fill = PatternFill(start_color=fill_color, fill_type="solid")

    for col in range(1, ws.max_column + 1):
        ws[f"{get_column_letter(col)}2"].font = Font(bold=True)
        ws[f"{get_column_letter(col)}2"].alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(output_excel)


def process_sales_zip(sales_folder: str) -> str:
    data_rows = []
    for file in glob.glob(sales_folder + '/*'):
        for f in glob.glob(file + '/*'):
            if f.lower().endswith(".pdf"):
                print(f"📄 Processing: {f}")
                try:
                    row = extract_sales_data(f)
                    data_rows.append(row)
                except Exception as e:
                    print(f"❌ Failed to process {f}: {e}")

    output_path = "monthly_sales_report.xlsx"
    if data_rows:
        df = pd.DataFrame(data_rows)
        write_to_excel_with_categories(df, output_path)
    return output_path
